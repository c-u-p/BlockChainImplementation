import tkinter
from tkinter import *
from tkinter import ttk
import hashlib

from openpyxl import load_workbook
wb = load_workbook('blockChainData.xlsx')
work_sheet = wb.active

root = Tk()
root.title("myBlockchain")
root.geometry('500x100')

L1 = Label(root ,text = "Prev Hash").pack()
c1 = Entry(root)
c1.pack()

L4 = Label(root, text="Data").pack()
d1 = Entry(root, state=DISABLED)
d1.pack()


def save1(event):
   
   var = d1.get()
   if var:
      res = hashlib.sha256(var.encode())
      st_res = res.hexdigest()
      work_sheet.append([st_res])
      wb.save('blockChainData.xlsx')
      d1.delete(0, END)
      top4= Toplevel(root)
      top4.geometry("500x50")
      top4.title("Block Created Successfully!")
      Label(top4, text= "congratulation! Your data is added in new block of blockchain.").pack()
      d1.configure(state=DISABLED)
      c1.configure(state=NORMAL)
      c1.delete(0, END)
      print(var)
      print(st_res)
      print('-------')   
   else:
      top= Toplevel(root)
      top.geometry("500x50")
      top.title("Value Warning!")
      Label(top, text= "Please, do not enter empty value!").pack()


def save0(event0):
   cval = c1.get()
   if cval:
      last_row = work_sheet.max_row
      prev_val = work_sheet.cell(column=1, row=last_row).value
      if prev_val == None:
         top2= Toplevel(root)
         top2.geometry("500x50")
         top2.title("Previous Hash Status!")
         Label(top2, text= "This is the First Block. So, no previous block exist! Please enter the data in first block of blockchain!").pack()
         c1.configure(state=DISABLED)
         d1.configure(state=NORMAL)
      elif cval != prev_val:
         result = hashlib.sha256(prev_val.encode())
         prev_hash = result.hexdigest()
         root.destroy()
      elif cval == prev_val:
         top3= Toplevel(root)
         top3.geometry("500x50")
         top3.title("Access Permitted!")
         Label(top3, text= "Please, enter the data in block of Blockchain!").pack()
         c1.configure(state=DISABLED)
         d1.configure(state=NORMAL)
   else:
      root.destroy()

c1.bind("<Return>", save0)
d1.bind("<Return>", save1)

root.mainloop()
